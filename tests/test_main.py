#!/usr/bin/env python3
"""
Tests unitaires pour main.py
Tests des fonctions utilitaires et de parsing.
"""

import pytest

from main import VideoOverlayAutomator


class TestVideoOverlayAutomator:
    """Tests pour la classe VideoOverlayAutomator."""

    @pytest.fixture
    def automator(self, tmp_path):
        """Fixture pour créer un automator avec fichiers temporaires."""
        # Créer des fichiers vides pour éviter les erreurs
        xml_file = tmp_path / "test.xml"
        xml_file.write_text(
            '<?xml version="1.0"?><xmeml version="5"><sequence><media><video><track></track></video></media></sequence></xmeml>')

        excel_file = tmp_path / "test.xlsx"
        # Créer un fichier Excel minimal avec openpyxl
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Set", "Point", "Set1", "Set2", "Jeux", "Points", "Commentaires"])
        ws.append([1, 1, None, None, "0/0", "0/0", ""])
        wb.save(excel_file)

        return VideoOverlayAutomator(
            xml_path=str(xml_file),
            excel_path=str(excel_file),
            video_folder=str(tmp_path)
        )

    def test_frames_to_seconds_zero(self, automator):
        """Test de conversion de 0 frames en secondes."""
        result = automator.frames_to_seconds(0)
        assert result == 0.0

    def test_frames_to_seconds_one_second(self, automator):
        """Test de conversion de ~60 frames (1 seconde en NTSC)."""
        # NTSC = 59.94 fps
        result = automator.frames_to_seconds(59.94)
        assert pytest.approx(result, rel=1e-2) == 1.0

    def test_frames_to_seconds_multiple_seconds(self, automator):
        """Test de conversion de plusieurs secondes."""
        # 300 frames ≈ 5 secondes à 59.94 fps
        result = automator.frames_to_seconds(300)
        assert pytest.approx(result, rel=1e-2) == 5.005

    def test_frames_to_seconds_decimal(self, automator):
        """Test de conversion avec frames décimales."""
        result = automator.frames_to_seconds(119.88)
        assert pytest.approx(result, rel=1e-2) == 2.0

    def test_format_time_seconds_only(self, automator):
        """Test de formatage du temps (secondes uniquement)."""
        result = automator.format_time(45)
        assert result == "45s"

    def test_format_time_minutes_and_seconds(self, automator):
        """Test de formatage du temps (minutes et secondes)."""
        result = automator.format_time(125)
        assert result == "2m05s"

    def test_format_time_hours_minutes_seconds(self, automator):
        """Test de formatage du temps (heures, minutes, secondes)."""
        result = automator.format_time(3665)
        assert result == "1h01m05s"

    def test_format_time_zero(self, automator):
        """Test de formatage du temps à zéro."""
        result = automator.format_time(0)
        assert result == "0s"

    def test_format_time_one_minute_exact(self, automator):
        """Test de formatage d'une minute exacte."""
        result = automator.format_time(60)
        assert result == "1m00s"

    def test_format_time_one_hour_exact(self, automator):
        """Test de formatage d'une heure exacte."""
        result = automator.format_time(3600)
        assert result == "1h00m00s"

    def test_format_time_multiple_hours(self, automator):
        """Test de formatage de plusieurs heures."""
        result = automator.format_time(7325)  # 2h02m05s
        assert result == "2h02m05s"

    def test_detect_gpu_encoder_returns_dict(self, automator):
        """Test que detect_gpu_encoder retourne un dictionnaire."""
        encoder = automator.detect_gpu_encoder()

        assert isinstance(encoder, dict)
        assert 'video_codec' in encoder
        assert 'preset' in encoder
        assert 'crf' in encoder
        assert 'extra_params' in encoder

    def test_detect_gpu_encoder_valid_codec(self, automator):
        """Test que detect_gpu_encoder retourne un codec valide."""
        encoder = automator.detect_gpu_encoder()

        valid_codecs = ['hevc_videotoolbox', 'hevc_nvenc', 'libx264']
        assert encoder['video_codec'] in valid_codecs

    def test_initialization_with_custom_teams(self, tmp_path):
        """Test de l'initialisation avec noms d'équipes personnalisés."""
        xml_file = tmp_path / "test.xml"
        xml_file.write_text(
            '<?xml version="1.0"?><xmeml version="5"><sequence><media><video><track></track></video></media></sequence></xmeml>')

        excel_file = tmp_path / "test.xlsx"
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Set", "Point", "Set1", "Set2", "Jeux", "Points", "Commentaires"])
        wb.save(excel_file)

        automator = VideoOverlayAutomator(
            xml_path=str(xml_file),
            excel_path=str(excel_file),
            video_folder=str(tmp_path),
            team1_names="TEAM A",
            team2_names="TEAM B"
        )

        assert automator.team1_names == "TEAM A"
        assert automator.team2_names == "TEAM B"

    def test_initialization_default_teams(self, tmp_path):
        """Test de l'initialisation avec noms d'équipes par défaut."""
        xml_file = tmp_path / "test.xml"
        xml_file.write_text(
            '<?xml version="1.0"?><xmeml version="5"><sequence><media><video><track></track></video></media></sequence></xmeml>')

        excel_file = tmp_path / "test.xlsx"
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Set", "Point", "Set1", "Set2", "Jeux", "Points", "Commentaires"])
        wb.save(excel_file)

        automator = VideoOverlayAutomator(
            xml_path=str(xml_file),
            excel_path=str(excel_file),
            video_folder=str(tmp_path)
        )

        assert automator.team1_names == "LÉO / YANNOUCK"
        assert automator.team2_names == "BILAL / PIERRE"

    def test_fps_is_ntsc(self, automator):
        """Test que le FPS est configuré en NTSC."""
        assert automator.fps == 59.94

    def test_parse_excel_basic(self, automator):
        """Test du parsing Excel basique."""
        scores = automator.parse_excel()

        assert isinstance(scores, list)
        assert len(scores) >= 1
        # Vérifier la structure du premier score
        if scores:
            assert 'set' in scores[0]
            assert 'point' in scores[0]
            assert 'jeux' in scores[0]
            assert 'points' in scores[0]

    def test_clips_initialization(self, automator):
        """Test que les clips sont initialisés comme liste vide."""
        assert isinstance(automator.clips, list)
        assert len(automator.clips) == 0

    def test_scores_initialization(self, automator):
        """Test que les scores sont initialisés comme liste vide."""
        assert isinstance(automator.scores, list)
        assert len(automator.scores) == 0
